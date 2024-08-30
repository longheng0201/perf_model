import os
#import inspect
import sys
import openpyxl
import json
import math
import graph_class
#############################no use tflite tool
#import tflite
#sys.path.append(r"/home/hxbu/tflite_parser/tflite/tflite/")
#from utils_name import opcode2name #from file import function
#sys.path.append(r"/home/hxbu/tflite_parser/tflite/tests/")
from parser_model import parser_model
# Example of parsing a TFLite model with `tflite` python package.
# Use this package, you can *import* the `tflite* package ONLY ONCE.
# Otherwise, you need to import every class when using them.
#sys.path.append("..")
#import test_util.basic

def test_net():
    with open('op.json', 'r') as input_op:
        op_data = json.load(input_op)
        op_data_ops = op_data['operators']
    with open('cfg.json', 'r') as input_cfg:
        cfg_data = json.load(input_cfg)
        cfg_data_bands = cfg_data['band']
        cfg_len        = len(cfg_data_bands)    
    with open('mem_cfg.json', 'r') as mem_cfg:
        cfg_mem_size = json.load(mem_cfg)
        cfg_mem_bands = cfg_mem_size['mem']
        mem_len        = len(cfg_mem_bands)    
    cur_dir = os.path.dirname(os.path.abspath(__file__))
    tflm_dir = os.path.abspath(cur_dir) #+ '/../assets/tests')
    tflm_name = input("请输入模型文件名tflite： ") #'gpt2-64-8bits.tflite'
    path = os.path.join(tflm_dir, tflm_name)
    layer_info =[]
    new_graph = graph_class.Graph()
    layer_info,new_graph = parser_model(path)
    layer_rows = len(layer_info)
    layer_cols = len(layer_info[0])
    workbook_path = tflm_name+"info.xlsx" #input("请输入excel路径： ")
    try:
        os.remove(workbook_path)
        #workbook = openpyxl.Workbook()
        workbook = openpyxl.load_workbook(workbook_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    ############# model #########################################################
    #cfg_type = input("请输入cfg")
    for cfg_mem_set in cfg_mem_bands :
        mem_band_type    =cfg_mem_set['type']
        vrf_size         =cfg_mem_set['vrf_size']
        pm_size          =cfg_mem_set['pm_size']
        lm_size          =cfg_mem_set['lm_size']
        gm_size          =cfg_mem_set['gm_size']
        ddr_size         =cfg_mem_set['ddr_size']
        risc_num           =cfg_mem_set['risc_num']
        core_num           =cfg_mem_set['core_num']
        cluster_num           =cfg_mem_set['cluster_num']
        pm_band          =cfg_mem_set['pm_band']
        lm_band          =cfg_mem_set['lm_band']
        gm_band          =cfg_mem_set['gm_band']
        ddr_band         =cfg_mem_set['ddr_band']
        mem_cfg =[pm_size,lm_size,gm_size,ddr_size,risc_num,core_num,cluster_num]
        new_graph.mark_treewalker_buffer(mem_cfg)
        new_graph.print_treewalker()
        print ("treewlker buffer done")
        for cfg_data_band in cfg_data_bands:
            #if(cfg_data_band['type'] == cfg_type):
            cfg_type    =cfg_data_band['type']
            vcube_w     =cfg_data_band['vcube_w']
            vcube_h     =cfg_data_band['vcube_h']
            vcube_c     =cfg_data_band['vcube_c']
            tcube_w     =cfg_data_band['tcube_w']
            tcube_h     =cfg_data_band['tcube_h']
            tcube_ic     =cfg_data_band['tcube_ic']
            tcube_oc     =cfg_data_band['tcube_oc']
            tpo          =cfg_data_band['tpo']
            vpo          =cfg_data_band['vpo']
            inner_oc     =cfg_data_band['inner_oc']
            vrf_tensor= cfg_data_band['vrf_tensor'],
            vrf_vector= cfg_data_band['vrf_vector'],
            vector_vrf= cfg_data_band['vector_vrf'],
            wgt_tensor= cfg_data_band['wgt_tensor'],
            tensor_vrf= cfg_data_band['tensor_vrf'],

            #########每层信息 10 :idx, layer_type,iw,ih,ic,ow,oh,oc,kw,kh,op cnt
            layer_cols =19
            layer_rows =len(layer_info)
            layer_array = [[0 for j in range(layer_cols)] for i in range(layer_rows)]
            ######## 统计layer op 信息  5:optype,mac, layer_num,tensor_y,weight size
            rows = 157
            cols = 5
            array = [[0 for j in range(cols)] for i in range(rows)]
            for i in range(rows):
                for j in range(cols):
                    array[i][j] = 0
    
            tensor_num=0
            tensor_cnt=0
            vector_num=0
            vector_cnt=0
            total_time =0
            total_rd_times =0
            total_wr_times =0
            idx=0
            for node in new_graph.valid_nodes:
                ###from 
                req_band=0
                optype_test,opcode,iw,ih,ic,ow,oh,oc,kw,kh,sw,sh,dw,dh = layer_info[node.idx][:layer_cols]
                optype_test = optype_test.upper()
                for tensor in node.input_tensor:
                    ibuf_loc = tensor.loc
                for tensor in node.output_tensor:
                    obuf_loc = tensor.loc
                wbuf_loc = node.wbuf_loc
                name_space ={
                    'ceil':math.ceil,
                    'tcube_w':tcube_w,
                    'tcube_h':tcube_h,
                    'tcube_ic':tcube_ic,
                    'tcube_oc':tcube_oc,
                    'vcube_w':vcube_w,
                    'vcube_h':vcube_h,
                    'vcube_c':vcube_c,
                    "tpo":tpo,
                    "vpo":vpo,
                    "inner_oc":inner_oc,
                    'vrf_tensor':vrf_tensor,
                    'vrf_vector':vrf_vector,
                    'vector_vrf':vector_vrf,
                    'wgt_tensor':wgt_tensor,
                    'tensor_vrf':tensor_vrf,
                    'iw':iw,
                    'ih':ih,
                    'ic':ic,
                    'ow':ow,
                    'oh':oh,
                    'oc':oc,
                    'kw':kw,
                    'kh':kh,
                    'req_band':req_band
             
                }
                print("name_space",name_space)  ##find op type
                for cfg_data_op in op_data_ops:
                    if(optype_test in cfg_data_op['type']):
                        req_band = eval(cfg_data_op['req_band'],name_space)
                        name_space['req_band'] = req_band
                        tensor_y = cfg_data_op['tensor_y']
                        op_cnt   = eval(cfg_data_op['cnt'],name_space)
                        para     = eval(cfg_data_op['para'],name_space)
                        wgt_rd_times = eval(cfg_data_op['wgt_rd_times'],name_space)
                        fm_rd_times = eval(cfg_data_op['fm_rd_times'],name_space)
                        fm_wr_times = eval(cfg_data_op['fm_wr_times'],name_space)
                        formula  = eval(cfg_data_op['formula'],name_space)
                print("name_space",name_space)  ##find op type
                #####  XXXB/GB =  ns
                pm_rd_time = (wgt_rd_times *(wbuf_loc==0) + fm_rd_times *(ibuf_loc==0))/pm_band
                lm_rd_time = (wgt_rd_times *(wbuf_loc==1) + fm_rd_times *(ibuf_loc==1))/lm_band
                gm_rd_time = (wgt_rd_times *(wbuf_loc==2) + fm_rd_times *(ibuf_loc==2))/gm_band
                ddr_rd_time = (wgt_rd_times *(wbuf_loc==3) + fm_rd_times *(ibuf_loc==3))/ddr_band
                pm_wr_time = ( fm_wr_times *(obuf_loc==0))/pm_band
                lm_wr_time = ( fm_wr_times *(obuf_loc==1))/lm_band
                gm_wr_time = ( fm_wr_times *(obuf_loc==2))/gm_band
                ddr_wr_time = ( fm_wr_times *(obuf_loc==3))/ddr_band
                
                rd_time     = max(pm_rd_time,lm_rd_time,gm_rd_time,ddr_rd_time)
                wr_time     = max(pm_wr_time,lm_wr_time,gm_wr_time,ddr_wr_time)
                op_time     = max(rd_time,wr_time,formula)
                print("\n  op,wr,rd,op time",rd_time,wr_time,formula)
                total_time = total_time +op_time 
                total_rd_times = total_rd_times+ wgt_rd_times+fm_rd_times
                total_wr_times = total_wr_times+ fm_wr_times
                #########cnt all op type 
                array[opcode][0] = f"{optype_test}"
                array[opcode][1] = array[opcode][1] + op_cnt
                array[opcode][2] = array[opcode][2] + 1
                array[opcode][3] = tensor_y
                array[opcode][4] = array[opcode][4]+para
                #####layer every idx
                layer_array[idx][0] = idx
                layer_array[idx][1] = f"{optype_test}  {opcode}"
                layer_array[idx][2] = iw
                layer_array[idx][3] = ih
                layer_array[idx][4] = ic
                layer_array[idx][5] = ow
                layer_array[idx][6] = oh
                layer_array[idx][7] = oc
                layer_array[idx][8] = kw
                layer_array[idx][9] = kh
                layer_array[idx][10] = op_cnt
                layer_array[idx][11] = formula
                layer_array[idx][12] = wgt_rd_times
                layer_array[idx][13] = fm_rd_times
                layer_array[idx][14] = fm_wr_times
                layer_array[idx][15] = rd_time
                layer_array[idx][16] = wr_time
                layer_array[idx][17] = para
                layer_array[idx][18] = op_time-formula
                if tensor_y:
                    tensor_cnt = tensor_cnt + op_cnt
                    tensor_num = tensor_num + 1
                else:
                    vector_cnt = vector_cnt + op_cnt
                    vector_num = vector_num + 1
                #print("kh begin_",kh,kw,ic,oc,ow,oh,"______",op_cnt,"__",tensor_y)
                idx= idx+1

            print("\n","all tensor_cnt",tensor_cnt,"allvector_cnt",vector_cnt)   
            print("\n","all tensor_num",tensor_num,"allvector_num",vector_num)   
            print("total time",total_time)
            print("total rd times",total_rd_times)
            #print("\n","move op num",move_cnt,"\n")
            ##################write to excel     ###########################################
            info_name = mem_band_type+cfg_type+"info"
            sheet = workbook.create_sheet(title=info_name)
            #sheet.title ="network_info"
            #for j in range(cols):
            sheet.cell(0+1,0+1).value = "optype"
            sheet.cell(0+1,1+1).value = "mac num"
            sheet.cell(0+1,2+1).value = "layer num"
            sheet.cell(0+1,3+1).value = "tensor yes"
            sheet.cell(0+1,4+1).value = "weight size" 
            sheet.cell(0+1,5+1).value = "total time" 
            sheet.cell(0+1,6+1).value = "total rd time" 
            sheet.cell(1+1,5+1).value =  total_time
            sheet.cell(1+1,6+1).value =  total_rd_times
            real_row =0
            real_layer =0
            for i in range(rows):
                if array[i][2] != 0:
                    real_row = real_row+1
                    real_layer = 1
                else:
                    real_row = real_row
                    real_layer = 0
                for j in range(cols):
                    if(real_layer==1):  
                        sheet.cell(row=real_row+1, column=j+1).value = array[i][j]
            layer_info_name = mem_band_type +cfg_type+"layer"
            sheet2 = workbook.create_sheet(title=layer_info_name)
            sheet2.cell(row=1, column=1).value = "layer idx"
            sheet2.cell(row=1, column=2).value = "layer type"
            sheet2.cell(row=1, column=3).value = "iwcoc_set"
            sheet2.cell(row=1, column=4).value = "ih"
            sheet2.cell(row=1, column=5).value = "ic"
            sheet2.cell(row=1, column=6).value = "ow"
            sheet2.cell(row=1, column=7).value = "oh"
            sheet2.cell(row=1, column=8).value = "oc"
            sheet2.cell(row=1, column=9).value = "kw"
            sheet2.cell(row=1, column=10).value = "kh"
            sheet2.cell(row=1, column=11).value = "mac_cnt"
            sheet2.cell(row=1, column=12).value = "time"
            sheet2.cell(row=1, column=13).value = "wgt_rd_times"
            sheet2.cell(row=1, column=14).value = "fm_rd_times"
            sheet2.cell(row=1, column=15).value = "fm_wr_times"
            sheet2.cell(row=1, column=16).value = "rd_time"
            sheet2.cell(row=1, column=17).value = "wr_time"
            sheet2.cell(row=1, column=18).value = "param"
            sheet2.cell(row=1, column=19).value = "buf_bound"
            for i in range(layer_rows):
                for j in range(layer_cols):
                    sheet2.cell(row=i+2, column=j+1).value = layer_array[i][j]
    
    workbook.save(workbook_path) 
# 保存Excel文件
    #out_name = tflm_name+"output"+cfg_type
   


if __name__ == '__main__':
    test_net()
